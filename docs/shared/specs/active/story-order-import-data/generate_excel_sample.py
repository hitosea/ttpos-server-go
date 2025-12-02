#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成订单数据导入 Excel 示例文件
根据 excel-format.md 定义的格式生成示例文件
"""

import sys
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要安装 openpyxl 库")
    print("请运行: pip3 install openpyxl")
    sys.exit(1)

def create_excel_sample():
    """创建 Excel 示例文件"""
    
    # 创建工作簿
    wb = Workbook()
    
    # ========== Sheet1: 订单基本信息 ==========
    ws1 = wb.active
    ws1.title = "订单基本信息"
    
    # 表头行
    headers1 = [
        "订单号", "下单时间", "订单状态", "订单类型", "用餐方式",
        "订单金额", "订单原价", "门店名称", "桌台名称", "会员编号",
        "收银员姓名", "就餐人数", "当班编号", "桌位编号", "商品金额",
        "服务费", "税费", "备注"
    ]
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    for col_idx, header in enumerate(headers1, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 示例数据
    sample_data1 = [
        ["ORD001", "2025-11-19 14:30:00", 1, 0, 0, 128.50, 150.00, "华莱士(北京店)", "A01", "", "张三", 2, "DUTY001", "SERIAL001", 120.00, 5.00, 3.50, "不要辣"],
        ["ORD002", "2025-11-19 15:00:00", 1, 1, 1, 45.00, 50.00, "华莱士(北京店)", "", "", "李四", 1, "", "", 40.00, 2.00, 3.00, ""],
        ["ORD003", "2025-11-19 16:00:00", 1, 2, 0, 88.00, 100.00, "华莱士(上海店)", "B02", "M001234", "王五", 1, "DUTY002", "", 80.00, 5.00, 3.00, "会员订单"],
    ]
    
    # 写入数据
    for row_idx, row_data in enumerate(sample_data1, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, (int, float)):
                cell.number_format = '0.00' if isinstance(value, float) else '0'
    
    # 调整列宽
    column_widths1 = [15, 20, 12, 12, 12, 12, 12, 20, 12, 12, 12, 12, 12, 12, 12, 10, 10, 15]
    for col_idx, width in enumerate(column_widths1, start=1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = width
    
    # ========== Sheet2: 订单明细 ==========
    ws2 = wb.create_sheet("订单明细")
    
    # 表头行
    headers2 = ["订单号", "商品名称", "数量", "单价", "小计", "备注"]
    
    # 写入表头
    for col_idx, header in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 示例数据
    sample_data2 = [
        ["ORD001", "香辣鸡腿堡", 2, 15.00, 30.00, ""],
        ["ORD001", "薯条", 1, 12.00, 12.00, ""],
        ["ORD001", "可乐", 2, 8.00, 16.00, "去冰"],
        ["ORD002", "鸡米花", 1, 20.00, 20.00, ""],
        ["ORD002", "奶茶", 1, 15.00, 15.00, ""],
        ["ORD003", "汉堡套餐", 1, 88.00, 88.00, ""],
    ]
    
    # 写入数据
    for row_idx, row_data in enumerate(sample_data2, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, (int, float)):
                cell.number_format = '0.00' if isinstance(value, float) else '0'
    
    # 调整列宽
    column_widths2 = [15, 20, 10, 12, 12, 15]
    for col_idx, width in enumerate(column_widths2, start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    
    # 保存文件
    output_file = "订单数据导入示例.xlsx"
    wb.save(output_file)
    print(f"✅ Excel 示例文件已生成: {output_file}")
    print(f"   - Sheet1: 订单基本信息 ({len(sample_data1)} 条示例数据)")
    print(f"   - Sheet2: 订单明细 ({len(sample_data2)} 条示例数据)")

if __name__ == "__main__":
    try:
        create_excel_sample()
    except Exception as e:
        import traceback
        print(f"❌ 生成失败: {e}")
        traceback.print_exc()
        sys.exit(1)

